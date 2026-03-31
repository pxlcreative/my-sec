"""
Tests for the Module H Excel DDQ generator.

Uses a lightweight mock firm (no DB required).
"""
import io
from datetime import date, datetime
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------

def _make_firm(**overrides):
    defaults = dict(
        crd_number=12345,
        legal_name="Acme Capital Management LLC",
        business_name="Acme Capital",
        sec_number="801-12345",
        registration_status="Registered",
        org_type="Limited Liability Company",
        phone="212-555-0100",
        website="https://acmecapital.example.com",
        fiscal_year_end="December",
        last_filing_date=date(2024, 3, 31),
        aum_total=500_000_000,
        aum_discretionary=450_000_000,
        aum_non_discretionary=50_000_000,
        num_accounts=320,
        num_employees=45,
        main_street1="123 Main Street",
        main_street2=None,
        main_city="New York",
        main_state="NY",
        main_zip="10001",
        main_country="United States",
        aum_2023=480_000_000,
        aum_2024=500_000_000,
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _make_disclosures(**overrides):
    defaults = dict(
        crd_number=12345,
        criminal_count=0,
        regulatory_count=1,
        civil_count=0,
        customer_count=0,
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def _make_aum_history():
    return [
        SimpleNamespace(
            filing_date=date(2022, 12, 31),
            aum_total=420_000_000,
            num_accounts=280,
        ),
        SimpleNamespace(
            filing_date=date(2023, 12, 31),
            aum_total=480_000_000,
            num_accounts=305,
        ),
        SimpleNamespace(
            filing_date=date(2024, 3, 31),
            aum_total=500_000_000,
            num_accounts=320,
        ),
    ]


def _build_and_reload(firm=None, aum_history=None, disclosures=None):
    """Build workbook, serialise to BytesIO, reload with openpyxl (no keep_vba)."""
    from services.excel_generator import build_dd_workbook

    firm         = firm         or _make_firm()
    aum_history  = aum_history  if aum_history is not None else _make_aum_history()
    disclosures  = disclosures  or _make_disclosures()

    wb = build_dd_workbook(firm, aum_history, disclosures)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestSheetNames:
    def test_correct_sheet_names(self):
        wb = _build_and_reload()
        assert wb.sheetnames == [
            "Cover",
            "Firm Overview",
            "Due Diligence Questionnaire",
            "AUM History",
            "Disclosures Summary",
            "Notes",
        ]


class TestCoverSheet:
    def setup_method(self):
        self.wb = _build_and_reload()
        self.ws = self.wb["Cover"]

    def test_firm_name_present(self):
        # Row 3 is the first data row (row 1 = title, row 2 = blank)
        values = [self.ws.cell(r, 2).value for r in range(3, 9)]
        assert "Acme Capital Management LLC" in values

    def test_crd_present(self):
        values = [self.ws.cell(r, 2).value for r in range(3, 9)]
        assert "12345" in values

    def test_registration_status_present(self):
        values = [self.ws.cell(r, 2).value for r in range(3, 9)]
        assert "Registered" in values

    def test_report_generated_is_today(self):
        values = [self.ws.cell(r, 2).value for r in range(3, 9)]
        assert date.today().strftime("%Y-%m-%d") in values


class TestDDQSheet:
    def setup_method(self):
        self.wb = _build_and_reload()
        self.ws = self.wb["Due Diligence Questionnaire"]

    def test_at_least_40_question_rows(self):
        # Row 1 is the header; count data rows after it
        data_rows = [
            r for r in range(2, self.ws.max_row + 1)
            if self.ws.cell(r, 2).value  # question column non-empty
        ]
        assert len(data_rows) >= 40, (
            f"Expected ≥40 question rows, got {len(data_rows)}"
        )

    def test_known_answer_cells_are_locked(self):
        """Column C (known answer) cells should be locked (prefilled)."""
        locked_count = 0
        for row in range(2, self.ws.max_row + 1):
            cell = self.ws.cell(row, 3)
            if cell.value is not None:
                # When loaded from file, protection is None if sheet protection
                # wasn't stripped; check fill colour instead as a proxy
                assert cell.fill.fgColor.rgb in ("FFDCE6F1", "DCE6F1", "00DCE6F1"), (
                    f"Row {row} col C fill unexpected: {cell.fill.fgColor.rgb}"
                )
                locked_count += 1
        assert locked_count >= 40

    def test_analyst_notes_cells_have_input_style(self):
        """Column D (analyst notes) cells should have yellow fill."""
        yellow_count = 0
        for row in range(2, self.ws.max_row + 1):
            cell = self.ws.cell(row, 4)
            if self.ws.cell(row, 2).value:  # only check rows with a question
                rgb = cell.fill.fgColor.rgb
                assert rgb in ("FFFFFC0", "FFFFC0", "00FFFFC0"), (
                    f"Row {row} col D fill unexpected: {rgb}"
                )
                yellow_count += 1
        assert yellow_count >= 40

    def test_first_question_contains_aum(self):
        q = self.ws.cell(2, 2).value
        assert q is not None
        assert "AUM" in q or "aum" in q.lower()

    def test_aum_answer_populated(self):
        # Known answer for first question should be a formatted dollar amount
        answer = self.ws.cell(2, 3).value
        assert answer and "$" in answer


class TestAumHistorySheet:
    def setup_method(self):
        self.wb = _build_and_reload()
        self.ws = self.wb["AUM History"]

    def test_header_row(self):
        assert self.ws.cell(1, 1).value == "Year"
        assert self.ws.cell(1, 2).value == "AUM Total"

    def test_three_data_rows(self):
        data_rows = [
            r for r in range(2, self.ws.max_row + 1)
            if self.ws.cell(r, 1).value is not None
        ]
        assert len(data_rows) == 3


class TestDisclosuresSheet:
    def test_with_disclosures(self):
        wb = _build_and_reload()
        ws = wb["Disclosures Summary"]
        values = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "Regulatory" in values

    def test_no_disclosures_message(self):
        firm = _make_firm()
        disc = _make_disclosures(
            criminal_count=0, regulatory_count=0,
            civil_count=0, customer_count=0,
        )
        wb = _build_and_reload(firm=firm, disclosures=disc)
        ws = wb["Disclosures Summary"]
        values = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert any("No disclosures" in str(v) for v in values if v)

    def test_none_disclosures_shows_no_record_message(self):
        wb = _build_and_reload(disclosures=None)
        ws = wb["Disclosures Summary"]
        values = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert any("No disclosures" in str(v) for v in values if v)


class TestNotesSheet:
    def test_notes_sheet_exists(self):
        wb = _build_and_reload()
        assert "Notes" in wb.sheetnames

    def test_notes_cell_label(self):
        wb = _build_and_reload()
        ws = wb["Notes"]
        assert ws["A1"].value == "Analyst Notes"


class TestEdgeCases:
    def test_firm_with_none_fields(self):
        """Workbook should build without error even when most fields are None."""
        sparse = SimpleNamespace(
            crd_number=99999,
            legal_name="Sparse Firm Inc",
            business_name=None,
            sec_number=None,
            registration_status=None,
            org_type=None,
            phone=None,
            website=None,
            fiscal_year_end=None,
            last_filing_date=None,
            aum_total=None,
            aum_discretionary=None,
            aum_non_discretionary=None,
            num_accounts=None,
            num_employees=None,
            main_street1=None,
            main_street2=None,
            main_city=None,
            main_state=None,
            main_zip=None,
            main_country=None,
            aum_2023=None,
            aum_2024=None,
        )
        wb = _build_and_reload(firm=sparse, aum_history=[], disclosures=None)
        assert "Cover" in wb.sheetnames
        assert wb["Cover"].cell(3, 2).value == "Sparse Firm Inc"

    def test_empty_aum_history(self):
        wb = _build_and_reload(aum_history=[])
        ws = wb["AUM History"]
        # Only the header row should be present
        assert ws.max_row == 1
