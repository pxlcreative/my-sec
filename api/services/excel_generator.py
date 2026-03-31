"""
Module H – On-demand Due Diligence Excel workbook generator.

Entry point: build_dd_workbook(firm, aum_history, disclosures) → openpyxl.Workbook
"""
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_FILL_HEADER    = PatternFill("solid", fgColor="003366")
_FILL_PREFILLED = PatternFill("solid", fgColor="DCE6F1")
_FILL_INPUT     = PatternFill("solid", fgColor="FFFFC0")
_FILL_SECTION   = PatternFill("solid", fgColor="BDD7EE")   # section sub-header

_FONT_HEADER    = Font(name="Arial", size=11, bold=True,  color="FFFFFF")
_FONT_PREFILLED = Font(name="Arial", size=10)
_FONT_INPUT     = Font(name="Arial", size=10)
_FONT_SECTION   = Font(name="Arial", size=10, bold=True)
_FONT_COVER_TITLE = Font(name="Arial", size=16, bold=True)
_FONT_LABEL     = Font(name="Arial", size=10, bold=True)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN_SIDE   = Side(style="thin", color="AAAAAA")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_PROT_LOCKED   = Protection(locked=True)
_PROT_UNLOCKED = Protection(locked=False)

_SHEET_PASSWORD = "readonly"


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _fmt_money(v) -> str:
    try:
        return f"${int(v):,.0f}" if v is not None else "N/A"
    except (TypeError, ValueError):
        return "N/A"


def _fmt_date(v) -> str:
    if v is None:
        return "N/A"
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _style_header(cell, align=None):
    cell.fill  = _FILL_HEADER
    cell.font  = _FONT_HEADER
    cell.alignment = align or _ALIGN_CENTER
    cell.protection = _PROT_LOCKED


def _style_section(cell):
    cell.fill  = _FILL_SECTION
    cell.font  = _FONT_SECTION
    cell.alignment = _ALIGN_LEFT
    cell.protection = _PROT_LOCKED


def _style_label(cell):
    cell.font  = _FONT_LABEL
    cell.alignment = _ALIGN_LEFT
    cell.protection = _PROT_LOCKED


def _style_prefilled(cell, value=None):
    if value is not None:
        cell.value = value
    cell.fill  = _FILL_PREFILLED
    cell.font  = _FONT_PREFILLED
    cell.alignment = _ALIGN_LEFT
    cell.border = _THIN_BORDER
    cell.protection = _PROT_LOCKED


def _style_input(cell, value=""):
    cell.value = value
    cell.fill  = _FILL_INPUT
    cell.font  = _FONT_INPUT
    cell.alignment = _ALIGN_LEFT
    cell.border = _THIN_BORDER
    cell.protection = _PROT_UNLOCKED


def _autosize_columns(ws: Worksheet, min_width: int = 10, max_width: int = 60):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val_len = len(str(cell.value or ""))
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _protect_sheet(ws: Worksheet):
    """Lock the sheet but allow selection and editing of unlocked cells."""
    ws.protection.sheet          = True
    ws.protection.password       = _SHEET_PASSWORD
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False


# ---------------------------------------------------------------------------
# Sheet 1 — Cover
# ---------------------------------------------------------------------------

def _build_cover(ws: Worksheet, firm):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45

    # Title
    ws.row_dimensions[1].height = 36
    title_cell = ws["A1"]
    title_cell.value = "Due Diligence Questionnaire"
    title_cell.font  = _FONT_COVER_TITLE
    title_cell.alignment = _ALIGN_CENTER
    ws.merge_cells("A1:B1")

    ws.append([])  # blank row

    rows = [
        ("Firm Legal Name",    firm.legal_name or "N/A"),
        ("CRD Number",         str(firm.crd_number)),
        ("SEC Number",         firm.sec_number or "N/A"),
        ("Report Generated",   date.today().strftime("%Y-%m-%d")),
        ("Data as of",         _fmt_date(firm.last_filing_date)),
        ("Registration Status", firm.registration_status or "N/A"),
    ]
    for label, value in rows:
        ws.append([label, value])
        row_n = ws.max_row
        _style_label(ws.cell(row_n, 1))
        _style_prefilled(ws.cell(row_n, 2))

    _protect_sheet(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — Firm Overview
# ---------------------------------------------------------------------------

def _section_header(ws, label: str, col_span: int = 2):
    ws.append([label])
    row_n = ws.max_row
    _style_section(ws.cell(row_n, 1))
    if col_span > 1:
        ws.merge_cells(
            start_row=row_n, start_column=1,
            end_row=row_n,   end_column=col_span,
        )


def _data_row(ws, label: str, value):
    ws.append([label, str(value) if value is not None else "N/A"])
    row_n = ws.max_row
    _style_label(ws.cell(row_n, 1))
    _style_prefilled(ws.cell(row_n, 2))


def _build_overview(ws: Worksheet, firm):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48

    _section_header(ws, "Basic Information")
    _data_row(ws, "Legal Name",    firm.legal_name)
    _data_row(ws, "Business Name", firm.business_name)
    _data_row(ws, "CRD Number",    firm.crd_number)
    _data_row(ws, "SEC Number",    firm.sec_number)
    _data_row(ws, "Org Type",      firm.org_type)
    _data_row(ws, "Phone",         firm.phone)
    _data_row(ws, "Website",       firm.website)

    ws.append([])
    _section_header(ws, "Business Operations")
    _data_row(ws, "Registration Status", firm.registration_status)
    _data_row(ws, "Fiscal Year End",     firm.fiscal_year_end)
    _data_row(ws, "Last Filing Date",    _fmt_date(firm.last_filing_date))

    ws.append([])
    _section_header(ws, "Address")
    _data_row(ws, "Street 1",  firm.main_street1)
    _data_row(ws, "Street 2",  firm.main_street2)
    _data_row(ws, "City",      firm.main_city)
    _data_row(ws, "State",     firm.main_state)
    _data_row(ws, "ZIP",       firm.main_zip)
    _data_row(ws, "Country",   firm.main_country)

    ws.append([])
    _section_header(ws, "AUM Summary")
    _data_row(ws, "Total AUM",              _fmt_money(firm.aum_total))
    _data_row(ws, "Discretionary AUM",      _fmt_money(firm.aum_discretionary))
    _data_row(ws, "Non-Discretionary AUM",  _fmt_money(firm.aum_non_discretionary))
    _data_row(ws, "Number of Accounts",     firm.num_accounts)
    _data_row(ws, "Number of Employees",    firm.num_employees)

    _protect_sheet(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — Due Diligence Questionnaire (40 questions)
# ---------------------------------------------------------------------------

def _dd_questions(firm, disclosures) -> list[tuple[str, str]]:
    """Return list of (question, known_answer) tuples."""
    total_disc = 0
    if disclosures:
        total_disc = (
            (disclosures.criminal_count    or 0)
            + (disclosures.regulatory_count or 0)
            + (disclosures.civil_count      or 0)
            + (disclosures.customer_count   or 0)
        )

    filing_year = (
        str(firm.last_filing_date.year)
        if firm.last_filing_date else "N/A"
    )

    aum_2023_str = _fmt_money(getattr(firm, "aum_2023", None))
    aum_2024_str = _fmt_money(getattr(firm, "aum_2024", None))

    return [
        # Core identification
        ("What is the firm's total AUM?",
         _fmt_money(firm.aum_total)),
        ("How many client accounts does the firm serve?",
         str(firm.num_accounts) if firm.num_accounts is not None else "N/A"),
        ("What is the firm's primary business name?",
         firm.business_name or firm.legal_name or "N/A"),
        ("Is the firm currently registered with the SEC?",
         firm.registration_status or "N/A"),
        ("Does the firm have disciplinary disclosures?",
         f"Yes – {total_disc} total" if total_disc > 0 else "No disclosures on record"),
        ("What types of clients does the firm serve?",
         "See ADV Item 5D"),
        ("What compensation arrangements does the firm use?",
         "See ADV Item 5E"),
        ("Does the firm have custody of client assets?",
         "See ADV Item 9"),
        ("What is the year of the most recent ADV filing?",
         filing_year),
        ("Does the firm have private fund clients?",
         "See ADV Item 7B"),
        # AUM detail
        ("What is the discretionary AUM?",
         _fmt_money(firm.aum_discretionary)),
        ("What is the non-discretionary AUM?",
         _fmt_money(firm.aum_non_discretionary)),
        ("What was the firm's AUM for 2023?",
         aum_2023_str),
        ("What was the firm's AUM for 2024?",
         aum_2024_str),
        ("Has AUM grown or declined over the past two years?",
         "See AUM History sheet"),
        # Employees & organisation
        ("How many employees does the firm have?",
         str(firm.num_employees) if firm.num_employees is not None else "N/A"),
        ("What is the firm's legal organisational type?",
         firm.org_type or "N/A"),
        ("What is the firm's fiscal year end?",
         firm.fiscal_year_end or "N/A"),
        ("What is the firm's CRD number?",
         str(firm.crd_number)),
        ("What is the firm's SEC file number?",
         firm.sec_number or "N/A"),
        # Registration & regulatory
        ("What states is the firm registered in?",
         "See ADV Part 1 Item 2"),
        ("Is the firm registered as an investment company?",
         "See ADV Item 2A"),
        ("Does the firm rely on any exemptions from registration?",
         "See ADV Item 2B"),
        ("Has the firm ever been subject to a regulatory action?",
         f"Regulatory disclosures: {disclosures.regulatory_count if disclosures else 0}"),
        ("Has the firm been subject to any criminal proceedings?",
         f"Criminal disclosures: {disclosures.criminal_count if disclosures else 0}"),
        ("Has the firm been subject to any civil proceedings?",
         f"Civil disclosures: {disclosures.civil_count if disclosures else 0}"),
        ("Have there been customer complaints or arbitrations?",
         f"Customer disclosures: {disclosures.customer_count if disclosures else 0}"),
        # Investment management
        ("What investment strategies does the firm employ?",
         "See ADV Item 8"),
        ("Does the firm manage wrap fee programs?",
         "See ADV Item 4"),
        ("Does the firm sponsor or advise private funds?",
         "See ADV Item 7B"),
        ("Does the firm use sub-advisers?",
         "See ADV Item 8F"),
        ("Does the firm trade on margin or use leverage?",
         "See ADV Item 8"),
        ("Does the firm employ derivatives strategies?",
         "See ADV Item 8"),
        # Client relationships
        ("What is the minimum account size?",
         "See ADV Item 5F"),
        ("Does the firm provide financial planning services?",
         "See ADV Item 5G"),
        ("Does the firm participate in wrap fee programs as sponsor?",
         "See ADV Item 4"),
        # Operations
        ("What is the firm's primary business address?",
         f"{firm.main_street1 or ''}, {firm.main_city or ''}, {firm.main_state or ''} {firm.main_zip or ''}".strip(", ")),
        ("What is the firm's phone number?",
         firm.phone or "N/A"),
        ("What is the firm's website?",
         firm.website or "N/A"),
        ("What is the date of the most recent ADV filing?",
         _fmt_date(firm.last_filing_date)),
    ]


def _build_ddq(ws: Worksheet, firm, disclosures):
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 38

    # Column headers
    headers = ["#", "Due Diligence Question", "Known Answer (from ADV)", "Analyst Notes"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        _style_header(ws.cell(header_row, col_idx))

    questions = _dd_questions(firm, disclosures)
    for i, (question, answer) in enumerate(questions, 1):
        ws.append([i, question, answer, ""])
        row_n = ws.max_row
        ws.row_dimensions[row_n].height = 28
        # #
        c_num = ws.cell(row_n, 1)
        c_num.fill  = _FILL_PREFILLED
        c_num.font  = _FONT_PREFILLED
        c_num.alignment = _ALIGN_CENTER
        c_num.border = _THIN_BORDER
        c_num.protection = _PROT_LOCKED
        # Question
        c_q = ws.cell(row_n, 2)
        c_q.fill  = _FILL_PREFILLED
        c_q.font  = _FONT_PREFILLED
        c_q.alignment = _ALIGN_LEFT
        c_q.border = _THIN_BORDER
        c_q.protection = _PROT_LOCKED
        # Known answer
        _style_prefilled(ws.cell(row_n, 3))
        # Analyst notes — unlocked
        _style_input(ws.cell(row_n, 4))

    _protect_sheet(ws)


# ---------------------------------------------------------------------------
# Sheet 4 — AUM History
# ---------------------------------------------------------------------------

def _build_aum_history(ws: Worksheet, aum_history: list):
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    headers = ["Year", "AUM Total", "Num Accounts", "Filing Date"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        _style_header(ws.cell(header_row, col_idx))

    for record in aum_history:
        # Support both ORM objects and dicts
        if hasattr(record, "filing_date"):
            year        = record.filing_date.year if record.filing_date else "N/A"
            aum_total   = record.aum_total
            num_accounts = record.num_accounts
            filing_date  = _fmt_date(record.filing_date)
        else:
            year        = record.get("year", "N/A")
            aum_total   = record.get("aum_total")
            num_accounts = record.get("num_accounts")
            filing_date  = record.get("filing_date", "N/A")

        ws.append([year, _fmt_money(aum_total), num_accounts, filing_date])
        row_n = ws.max_row
        for col_idx in range(1, 5):
            _style_prefilled(ws.cell(row_n, col_idx))

    _protect_sheet(ws)


# ---------------------------------------------------------------------------
# Sheet 5 — Disclosures Summary
# ---------------------------------------------------------------------------

def _build_disclosures(ws: Worksheet, disclosures):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16

    headers = ["Disclosure Type", "Count"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        _style_header(ws.cell(header_row, col_idx))

    if disclosures is None or (
        (disclosures.criminal_count    or 0) == 0
        and (disclosures.regulatory_count or 0) == 0
        and (disclosures.civil_count      or 0) == 0
        and (disclosures.customer_count   or 0) == 0
    ):
        ws.append(["No disclosures on record", ""])
        row_n = ws.max_row
        _style_prefilled(ws.cell(row_n, 1))
        _style_prefilled(ws.cell(row_n, 2))
    else:
        disc_rows = [
            ("Criminal",   disclosures.criminal_count    or 0),
            ("Regulatory", disclosures.regulatory_count  or 0),
            ("Civil",      disclosures.civil_count       or 0),
            ("Customer",   disclosures.customer_count    or 0),
        ]
        for label, count in disc_rows:
            ws.append([label, count])
            row_n = ws.max_row
            _style_prefilled(ws.cell(row_n, 1))
            _style_prefilled(ws.cell(row_n, 2))

    _protect_sheet(ws)


# ---------------------------------------------------------------------------
# Sheet 6 — Notes
# ---------------------------------------------------------------------------

def _build_notes(ws: Worksheet):
    ws.merge_cells("A1:F30")
    top_cell = ws["A1"]
    top_cell.value = "Analyst Notes"
    top_cell.font  = Font(name="Arial", size=10, bold=True, color="888888")
    top_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    top_cell.fill  = _FILL_INPUT
    top_cell.border = _THIN_BORDER
    top_cell.protection = _PROT_UNLOCKED

    # Notes sheet is intentionally NOT protected so the whole area is editable


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_dd_workbook(firm, aum_history: list, disclosures) -> Workbook:
    """
    Build and return a formatted openpyxl Workbook for the given firm.

    Args:
        firm:        ORM Firm object (or any object with the expected attributes)
        aum_history: list of FirmAumHistory ORM objects (or dicts with year/aum_total/…)
        disclosures: FirmDisclosuresSummary ORM object, or None
    """
    wb = Workbook()

    # Rename default sheet and build each sheet
    ws1 = wb.active
    ws1.title = "Cover"
    _build_cover(ws1, firm)

    ws2 = wb.create_sheet("Firm Overview")
    _build_overview(ws2, firm)

    ws3 = wb.create_sheet("Due Diligence Questionnaire")
    _build_ddq(ws3, firm, disclosures)

    ws4 = wb.create_sheet("AUM History")
    _build_aum_history(ws4, aum_history)

    ws5 = wb.create_sheet("Disclosures Summary")
    _build_disclosures(ws5, disclosures)

    ws6 = wb.create_sheet("Notes")
    _build_notes(ws6)

    # Auto-size all sheets (Notes excluded — it's a merged block)
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        _autosize_columns(ws)

    return wb
