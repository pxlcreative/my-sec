"""
Business logic for the questionnaire system.
"""
from __future__ import annotations

import io
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from models.aum import FirmAumHistory
from models.disclosures import FirmDisclosuresSummary
from models.firm import Firm
from models.questionnaire import (
    QuestionnaireQuestion,
    QuestionnaireResponse,
    QuestionnaireTemplate,
)
from services.questionnaire_resolver import resolve_answer, resolve_fields


# ---------------------------------------------------------------------------
# Template CRUD
# ---------------------------------------------------------------------------

def get_templates(db: Session) -> list[QuestionnaireTemplate]:
    return list(db.scalars(
        select(QuestionnaireTemplate)
        .options(selectinload(QuestionnaireTemplate.questions))
        .order_by(QuestionnaireTemplate.name)
    ).all())


def get_template(template_id: int, db: Session) -> QuestionnaireTemplate | None:
    return db.scalar(
        select(QuestionnaireTemplate)
        .options(selectinload(QuestionnaireTemplate.questions))
        .where(QuestionnaireTemplate.id == template_id)
    )


def create_template(data: dict[str, Any], db: Session) -> QuestionnaireTemplate:
    tmpl = QuestionnaireTemplate(
        name=data["name"],
        description=data.get("description"),
        style_type=data.get("style_type", "custom"),
    )
    db.add(tmpl)
    db.commit()
    db.refresh(tmpl)
    return tmpl


def update_template(template_id: int, data: dict[str, Any], db: Session) -> QuestionnaireTemplate | None:
    tmpl = db.get(QuestionnaireTemplate, template_id)
    if tmpl is None:
        return None
    for field in ("name", "description", "style_type"):
        if field in data:
            setattr(tmpl, field, data[field])
    db.commit()
    db.refresh(tmpl)
    return tmpl


def delete_template(template_id: int, db: Session) -> bool:
    """Returns False if template not found."""
    tmpl = db.get(QuestionnaireTemplate, template_id)
    if tmpl is None:
        return False
    db.delete(tmpl)
    db.commit()
    return True


def has_responses(template_id: int, db: Session) -> bool:
    return db.scalar(
        select(QuestionnaireResponse.id)
        .where(QuestionnaireResponse.template_id == template_id)
        .limit(1)
    ) is not None


# ---------------------------------------------------------------------------
# Question CRUD
# ---------------------------------------------------------------------------

def add_question(template_id: int, data: dict[str, Any], db: Session) -> QuestionnaireQuestion | None:
    tmpl = db.get(QuestionnaireTemplate, template_id)
    if tmpl is None:
        return None
    # Place at end of current questions
    max_order = db.scalar(
        select(QuestionnaireQuestion.order_index)
        .where(QuestionnaireQuestion.template_id == template_id)
        .order_by(QuestionnaireQuestion.order_index.desc())
        .limit(1)
    )
    next_order = (max_order or 0) + 1

    q = QuestionnaireQuestion(
        template_id=template_id,
        section=data.get("section", "General"),
        order_index=data.get("order_index", next_order),
        question_text=data["question_text"],
        answer_field_path=data.get("answer_field_path"),
        answer_hint=data.get("answer_hint"),
        notes_enabled=data.get("notes_enabled", True),
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    return q


def update_question(question_id: int, data: dict[str, Any], db: Session) -> QuestionnaireQuestion | None:
    q = db.get(QuestionnaireQuestion, question_id)
    if q is None:
        return None
    for field in ("section", "order_index", "question_text", "answer_field_path", "answer_hint", "notes_enabled"):
        if field in data:
            setattr(q, field, data[field])
    db.commit()
    db.refresh(q)
    return q


def delete_question(question_id: int, db: Session) -> bool:
    q = db.get(QuestionnaireQuestion, question_id)
    if q is None:
        return False
    db.delete(q)
    db.commit()
    return True


def reorder_questions(template_id: int, ordered_ids: list[int], db: Session) -> bool:
    questions = {
        q.id: q
        for q in db.scalars(
            select(QuestionnaireQuestion).where(QuestionnaireQuestion.template_id == template_id)
        ).all()
    }
    for idx, q_id in enumerate(ordered_ids):
        if q_id in questions:
            questions[q_id].order_index = idx
    db.commit()
    return True


# ---------------------------------------------------------------------------
# Response management
# ---------------------------------------------------------------------------

def _load_firm_context(crd: int, db: Session):
    """Load firm, disclosures, and aum_history for field resolution."""
    firm = db.get(Firm, crd)
    disclosures = db.get(FirmDisclosuresSummary, crd)
    aum_history = list(db.scalars(
        select(FirmAumHistory)
        .where(FirmAumHistory.crd_number == crd)
        .order_by(FirmAumHistory.filing_date.desc())
    ).all())
    return firm, disclosures, aum_history


def _auto_populate_answers(
    questions: list[QuestionnaireQuestion],
    firm,
    disclosures,
    aum_history: list,
) -> dict[str, str]:
    """Resolve all field-path answers. Returns {str(question_id): answer_str}."""
    resolved = resolve_fields(firm, disclosures, aum_history)
    answers: dict[str, str] = {}
    for q in questions:
        if q.answer_field_path:
            answers[str(q.id)] = resolve_answer(q.answer_field_path, resolved, firm)
    return answers


def get_or_create_response(template_id: int, crd: int, db: Session) -> QuestionnaireResponse | None:
    """
    Return the existing response for this template+firm, refreshing auto-resolved answers.
    Creates a new response if none exists.
    Manually edited answers (questions with no field_path) are preserved.
    """
    template = get_template(template_id, db)
    if template is None:
        return None

    firm, disclosures, aum_history = _load_firm_context(crd, db)
    if firm is None:
        return None

    auto_answers = _auto_populate_answers(template.questions, firm, disclosures, aum_history)

    response = db.scalar(
        select(QuestionnaireResponse)
        .where(
            QuestionnaireResponse.template_id == template_id,
            QuestionnaireResponse.crd_number == crd,
        )
    )

    if response is None:
        response = QuestionnaireResponse(
            template_id=template_id,
            crd_number=crd,
            answers=auto_answers,
            analyst_notes={},
            ai_suggested={},
            status="draft",
        )
        db.add(response)
    else:
        # Merge: auto-populated answers overwrite field_path answers;
        # manual edits for questions without a field_path are preserved.
        existing = dict(response.answers or {})
        field_path_question_ids = {
            str(q.id) for q in template.questions if q.answer_field_path
        }
        # Remove stale auto-answers, keep manual ones
        merged = {k: v for k, v in existing.items() if k not in field_path_question_ids}
        merged.update(auto_answers)
        response.answers = merged

    db.commit()
    db.refresh(response)
    return response


def regenerate_response(template_id: int, crd: int, db: Session) -> QuestionnaireResponse | None:
    """Force-refresh all auto-resolved answers, preserving analyst_notes."""
    template = get_template(template_id, db)
    if template is None:
        return None

    firm, disclosures, aum_history = _load_firm_context(crd, db)
    if firm is None:
        return None

    auto_answers = _auto_populate_answers(template.questions, firm, disclosures, aum_history)

    response = db.scalar(
        select(QuestionnaireResponse)
        .where(
            QuestionnaireResponse.template_id == template_id,
            QuestionnaireResponse.crd_number == crd,
        )
    )

    if response is None:
        response = QuestionnaireResponse(
            template_id=template_id,
            crd_number=crd,
            answers=auto_answers,
            analyst_notes={},
            ai_suggested={},
            status="draft",
        )
        db.add(response)
    else:
        response.answers = auto_answers
        response.ai_suggested = {}

    db.commit()
    db.refresh(response)
    return response


def update_response(
    template_id: int,
    crd: int,
    answers: dict[str, str],
    analyst_notes: dict[str, str],
    db: Session,
) -> QuestionnaireResponse | None:
    response = db.scalar(
        select(QuestionnaireResponse)
        .where(
            QuestionnaireResponse.template_id == template_id,
            QuestionnaireResponse.crd_number == crd,
        )
    )
    if response is None:
        return None

    if answers:
        merged = dict(response.answers or {})
        merged.update(answers)
        response.answers = merged

    if analyst_notes:
        merged_notes = dict(response.analyst_notes or {})
        merged_notes.update(analyst_notes)
        response.analyst_notes = merged_notes

    db.commit()
    db.refresh(response)
    return response


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_questionnaire_xlsx(
    response: QuestionnaireResponse,
    template: QuestionnaireTemplate,
    questions: list[QuestionnaireQuestion],
    firm,
) -> bytes:
    """Build and return Excel workbook bytes for the questionnaire response."""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter

    from datetime import date as dt_date

    # Reuse styles from excel_generator
    FILL_HEADER    = PatternFill("solid", fgColor="003366")
    FILL_PREFILLED = PatternFill("solid", fgColor="DCE6F1")
    FILL_INPUT     = PatternFill("solid", fgColor="FFFFC0")
    FILL_SECTION   = PatternFill("solid", fgColor="BDD7EE")
    FILL_AI        = PatternFill("solid", fgColor="E8F5E9")

    FONT_HEADER    = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    FONT_PREFILLED = Font(name="Arial", size=10)
    FONT_INPUT     = Font(name="Arial", size=10)
    FONT_SECTION   = Font(name="Arial", size=10, bold=True)
    FONT_COVER     = Font(name="Arial", size=16, bold=True)
    FONT_LABEL     = Font(name="Arial", size=10, bold=True)

    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    THIN_SIDE   = Side(style="thin", color="AAAAAA")
    THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    PROT_LOCKED   = Protection(locked=True)
    PROT_UNLOCKED = Protection(locked=False)

    has_ai = bool(response.ai_suggested)
    answers = response.answers or {}
    notes   = response.analyst_notes or {}
    ai_sugg = response.ai_suggested or {}

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Cover
    # ------------------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions["A"].width = 28
    ws_cover.column_dimensions["B"].width = 48

    ws_cover.row_dimensions[1].height = 36
    title_cell = ws_cover["A1"]
    title_cell.value = template.name
    title_cell.font = FONT_COVER
    title_cell.alignment = ALIGN_CENTER
    ws_cover.merge_cells("A1:B1")

    ws_cover.append([])

    cover_rows = [
        ("Firm Legal Name",    firm.legal_name or "N/A"),
        ("CRD Number",         str(firm.crd_number)),
        ("Registration Status", firm.registration_status or "N/A"),
        ("Template Style",     template.style_type),
        ("Status",             response.status.capitalize()),
        ("Report Generated",   dt_date.today().strftime("%Y-%m-%d")),
    ]
    for label, value in cover_rows:
        ws_cover.append([label, value])
        row_n = ws_cover.max_row
        lc = ws_cover.cell(row_n, 1)
        lc.font = FONT_LABEL
        lc.alignment = ALIGN_LEFT
        lc.protection = PROT_LOCKED
        vc = ws_cover.cell(row_n, 2)
        vc.value = value
        vc.fill = FILL_PREFILLED
        vc.font = FONT_PREFILLED
        vc.alignment = ALIGN_LEFT
        vc.border = THIN_BORDER
        vc.protection = PROT_LOCKED

    ws_cover.protection.sheet = True
    ws_cover.protection.password = "readonly"
    ws_cover.protection.selectLockedCells = False
    ws_cover.protection.selectUnlockedCells = False

    # ------------------------------------------------------------------
    # Sheet 2: Questionnaire (one sheet, grouped by section)
    # ------------------------------------------------------------------
    ws_q = wb.create_sheet("Questionnaire")
    ws_q.column_dimensions["A"].width = 6
    ws_q.column_dimensions["B"].width = 52
    ws_q.column_dimensions["C"].width = 38
    ws_q.column_dimensions["D"].width = 38
    if has_ai:
        ws_q.column_dimensions["E"].width = 38

    # Header row
    headers = ["#", "Question", "Answer", "Analyst Notes"]
    if has_ai:
        headers.append("AI Suggestion")
    ws_q.append(headers)
    hrow = ws_q.max_row
    for ci, _ in enumerate(headers, 1):
        c = ws_q.cell(hrow, ci)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.protection = PROT_LOCKED

    # Group questions by section
    sections: dict[str, list[QuestionnaireQuestion]] = {}
    for q in sorted(questions, key=lambda x: (x.section, x.order_index)):
        sections.setdefault(q.section, []).append(q)

    q_num = 0
    for section_name, section_qs in sections.items():
        # Section header
        ws_q.append([section_name])
        srow = ws_q.max_row
        sc = ws_q.cell(srow, 1)
        sc.fill = FILL_SECTION
        sc.font = FONT_SECTION
        sc.alignment = ALIGN_LEFT
        sc.protection = PROT_LOCKED
        ncols = 5 if has_ai else 4
        ws_q.merge_cells(
            start_row=srow, start_column=1,
            end_row=srow,   end_column=ncols,
        )

        for q in section_qs:
            q_num += 1
            answer  = answers.get(str(q.id), "")
            note    = notes.get(str(q.id), "")
            ai_hint = ai_sugg.get(str(q.id), "") if has_ai else ""
            has_field = bool(q.answer_field_path)

            row_data = [q_num, q.question_text, answer, note]
            if has_ai:
                row_data.append(ai_hint)
            ws_q.append(row_data)
            rn = ws_q.max_row
            ws_q.row_dimensions[rn].height = 32

            # # col
            c_num = ws_q.cell(rn, 1)
            c_num.fill = FILL_PREFILLED
            c_num.font = FONT_PREFILLED
            c_num.alignment = ALIGN_CENTER
            c_num.border = THIN_BORDER
            c_num.protection = PROT_LOCKED

            # Question col
            c_q = ws_q.cell(rn, 2)
            c_q.fill = FILL_PREFILLED
            c_q.font = FONT_PREFILLED
            c_q.alignment = ALIGN_LEFT
            c_q.border = THIN_BORDER
            c_q.protection = PROT_LOCKED

            # Answer col — blue if auto-populated, yellow if blank/manual
            c_a = ws_q.cell(rn, 3)
            c_a.border = THIN_BORDER
            c_a.alignment = ALIGN_LEFT
            if has_field and answer:
                c_a.fill = FILL_PREFILLED
                c_a.font = FONT_PREFILLED
                c_a.protection = PROT_LOCKED
            else:
                c_a.fill = FILL_INPUT
                c_a.font = FONT_INPUT
                c_a.protection = PROT_UNLOCKED

            # Notes col
            if q.notes_enabled:
                c_n = ws_q.cell(rn, 4)
                c_n.fill = FILL_INPUT
                c_n.font = FONT_INPUT
                c_n.alignment = ALIGN_LEFT
                c_n.border = THIN_BORDER
                c_n.protection = PROT_UNLOCKED
            else:
                c_n = ws_q.cell(rn, 4)
                c_n.fill = FILL_PREFILLED
                c_n.font = FONT_PREFILLED
                c_n.alignment = ALIGN_LEFT
                c_n.border = THIN_BORDER
                c_n.protection = PROT_LOCKED

            # AI col
            if has_ai:
                c_ai = ws_q.cell(rn, 5)
                c_ai.fill = FILL_AI
                c_ai.font = FONT_PREFILLED
                c_ai.alignment = ALIGN_LEFT
                c_ai.border = THIN_BORDER
                c_ai.protection = PROT_LOCKED

    ws_q.protection.sheet = True
    ws_q.protection.password = "readonly"
    ws_q.protection.selectLockedCells = False
    ws_q.protection.selectUnlockedCells = False

    # ------------------------------------------------------------------
    # Sheet 3: Notes
    # ------------------------------------------------------------------
    ws_notes = wb.create_sheet("Notes")
    ws_notes.merge_cells("A1:F30")
    top_cell = ws_notes["A1"]
    top_cell.value = "Analyst Notes"
    top_cell.font = Font(name="Arial", size=10, bold=True, color="888888")
    top_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    top_cell.fill = FILL_INPUT
    top_cell.border = THIN_BORDER
    top_cell.protection = PROT_UNLOCKED

    # Auto-size cover + questionnaire
    for ws in [ws_cover, ws_q]:
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 10
            for cell in col_cells:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
