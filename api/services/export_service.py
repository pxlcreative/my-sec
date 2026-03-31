"""
Export service: query building, formatting (CSV / JSON / XLSX), and job orchestration.
"""
import csv
import io
import json
import logging
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from config import settings
from models.export_job import ExportJob
from models.firm import Firm
from models.platform import FirmPlatform, PlatformDefinition

log = logging.getLogger(__name__)

EXPORTS_DIR = Path(settings.data_dir) / "exports"
EXPORT_TTL_HOURS = 24
SYNC_ROW_LIMIT = 500

# Columns always included in exports (in order)
DEFAULT_FIELDS = [
    "crd_number", "legal_name", "business_name",
    "main_street1", "main_city", "main_state", "main_zip",
    "aum_total", "registration_status", "last_filing_date",
    "platforms",
]

# Human-readable CSV/XLSX headers matching DEFAULT_FIELDS
FIELD_HEADERS = {
    "crd_number":          "CRD_NUMBER",
    "legal_name":          "LEGAL_NAME",
    "business_name":       "BUSINESS_NAME",
    "main_street1":        "STREET",
    "main_city":           "CITY",
    "main_state":          "STATE",
    "main_zip":            "ZIP",
    "aum_total":           "AUM_TOTAL",
    "registration_status": "REGISTRATION_STATUS",
    "last_filing_date":    "LAST_FILING_DATE",
    "platforms":           "PLATFORMS",
    # Optional extras
    "sec_number":          "SEC_NUMBER",
    "main_street2":        "STREET2",
    "main_country":        "COUNTRY",
    "phone":               "PHONE",
    "website":             "WEBSITE",
    "org_type":            "ORG_TYPE",
    "fiscal_year_end":     "FISCAL_YEAR_END",
    "aum_discretionary":   "AUM_DISCRETIONARY",
    "aum_non_discretionary": "AUM_NON_DISCRETIONARY",
    "num_accounts":        "NUM_ACCOUNTS",
    "num_employees":       "NUM_EMPLOYEES",
}

_HEADER_FILL = PatternFill("solid", fgColor="003366")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


# ---------------------------------------------------------------------------
# 1a. Query builder
# ---------------------------------------------------------------------------

def build_export_query(filter_criteria: dict | None, crd_list: list[int] | None):
    """Return a SQLAlchemy select statement for firms matching the given filters."""
    stmt = select(Firm)

    if crd_list:
        stmt = stmt.where(Firm.crd_number.in_(crd_list))
        return stmt  # CRD list is authoritative; ignore other filters

    if filter_criteria:
        fc = filter_criteria
        if fc.get("registration_status"):
            stmt = stmt.where(Firm.registration_status == fc["registration_status"])
        if fc.get("aum_min") is not None:
            stmt = stmt.where(Firm.aum_total >= fc["aum_min"])
        if fc.get("aum_max") is not None:
            stmt = stmt.where(Firm.aum_total <= fc["aum_max"])
        if fc.get("states"):
            stmt = stmt.where(Firm.main_state.in_([s.upper() for s in fc["states"]]))
        if fc.get("platform_ids"):
            stmt = stmt.join(
                FirmPlatform,
                (FirmPlatform.crd_number == Firm.crd_number)
                & (FirmPlatform.platform_id.in_(fc["platform_ids"])),
            ).distinct()

    return stmt.order_by(Firm.legal_name)


def _count_query(stmt, session: Session) -> int:
    from sqlalchemy import func
    return session.scalar(select(func.count()).select_from(stmt.subquery())) or 0


def _fetch_rows(stmt, session: Session) -> list[dict]:
    """Fetch firms + their platform names as a list of plain dicts."""
    firms = list(session.scalars(stmt).all())
    if not firms:
        return []
    crds = [f.crd_number for f in firms]
    # Batch load platforms
    plat_rows = session.execute(
        select(FirmPlatform.crd_number, PlatformDefinition.name)
        .join(PlatformDefinition, FirmPlatform.platform_id == PlatformDefinition.id)
        .where(FirmPlatform.crd_number.in_(crds))
    ).all()
    plat_map: dict[int, list[str]] = {}
    for crd, name in plat_rows:
        plat_map.setdefault(crd, []).append(name)

    rows = []
    for f in firms:
        rows.append({
            "crd_number":          f.crd_number,
            "legal_name":          f.legal_name,
            "business_name":       f.business_name,
            "main_street1":        f.main_street1,
            "main_street2":        f.main_street2,
            "main_city":           f.main_city,
            "main_state":          f.main_state,
            "main_zip":            f.main_zip,
            "main_country":        f.main_country,
            "aum_total":           f.aum_total,
            "aum_discretionary":   f.aum_discretionary,
            "aum_non_discretionary": f.aum_non_discretionary,
            "num_accounts":        f.num_accounts,
            "num_employees":       f.num_employees,
            "registration_status": f.registration_status,
            "last_filing_date":    str(f.last_filing_date) if f.last_filing_date else None,
            "sec_number":          f.sec_number,
            "org_type":            f.org_type,
            "fiscal_year_end":     f.fiscal_year_end,
            "phone":               f.phone,
            "website":             f.website,
            "platforms":           ", ".join(plat_map.get(f.crd_number, [])),
        })
    return rows


def _active_fields(field_selection: list[str] | None) -> list[str]:
    if not field_selection:
        return DEFAULT_FIELDS
    extras = [f for f in field_selection if f not in DEFAULT_FIELDS and f in FIELD_HEADERS]
    return DEFAULT_FIELDS + extras


# ---------------------------------------------------------------------------
# 1b. CSV
# ---------------------------------------------------------------------------

def export_to_csv(rows: list[dict], field_selection: list[str] | None) -> bytes:
    fields = _active_fields(field_selection)
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=[FIELD_HEADERS.get(f, f.upper()) for f in fields],
        extrasaction="ignore",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({FIELD_HEADERS.get(f, f.upper()): row.get(f, "") for f in fields})
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# 1c. JSON
# ---------------------------------------------------------------------------

def export_to_json(rows: list[dict], field_selection: list[str] | None) -> bytes:
    fields = _active_fields(field_selection)
    output = [{f: row.get(f) for f in fields} for row in rows]
    return json.dumps(output, default=str, indent=2).encode("utf-8")


# ---------------------------------------------------------------------------
# 1d. XLSX
# ---------------------------------------------------------------------------

def export_to_xlsx(rows: list[dict], field_selection: list[str] | None) -> bytes:
    fields = _active_fields(field_selection)
    headers = [FIELD_HEADERS.get(f, f.upper()) for f in fields]

    wb = openpyxl.Workbook()

    # ------ Sheet 1: Firm data ------
    ws1 = wb.active
    ws1.title = "Firms"
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws1.append([row.get(f, "") for f in fields])

    # Auto-size columns
    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws1.cell(r, col_idx).value or "")) for r in range(1, ws1.max_row + 1)),
            default=8,
        )
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ------ Sheet 2: Platform tags ------
    ws2 = wb.create_sheet("Platform Tags")
    tag_headers = ["CRD_NUMBER", "LEGAL_NAME", "PLATFORM", "TAGGED_AT", "TAGGED_BY", "NOTES"]
    ws2.append(tag_headers)
    for cell in ws2[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Flatten platforms back out (stored as CSV string in row["platforms"])
    for row in rows:
        for plat in (row.get("platforms") or "").split(", "):
            if plat:
                ws2.append([
                    row["crd_number"],
                    row["legal_name"],
                    plat,
                    "", "", "",  # tagged_at / tagged_by / notes not loaded here
                ])

    for col_idx in range(1, len(tag_headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws2.cell(r, col_idx).value or "")) for r in range(1, ws2.max_row + 1)),
            default=8,
        )
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Format dispatch
# ---------------------------------------------------------------------------

_MIME_TYPES = {
    "csv":  ("text/csv", "csv"),
    "json": ("application/json", "json"),
    "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
}

_FORMATTERS = {
    "csv":  export_to_csv,
    "json": export_to_json,
    "xlsx": export_to_xlsx,
}


def format_rows(rows: list[dict], fmt: str, field_selection: list[str] | None) -> bytes:
    formatter = _FORMATTERS.get(fmt)
    if formatter is None:
        raise ValueError(f"Unknown format: {fmt!r}")
    return formatter(rows, field_selection)


def mime_type(fmt: str) -> str:
    return _MIME_TYPES[fmt][0]


def file_extension(fmt: str) -> str:
    return _MIME_TYPES[fmt][1]


# ---------------------------------------------------------------------------
# 1e. Celery task helper (called by task, not FastAPI directly)
# ---------------------------------------------------------------------------

def run_export_job(job_id: str, session: Session) -> None:
    """
    Core logic for the async export Celery task.
    Loads ExportJob → runs query → formats → writes file → updates job.
    """
    job: ExportJob | None = session.get(ExportJob, uuid.UUID(job_id))
    if job is None:
        raise ValueError(f"ExportJob {job_id} not found")

    job.status = "running"
    session.commit()

    try:
        stmt = build_export_query(job.filter_criteria, job.crd_list)
        rows = _fetch_rows(stmt, session)

        field_sel: list[str] | None = None
        if job.field_selection and isinstance(job.field_selection, dict):
            field_sel = job.field_selection.get("fields")
        elif job.field_selection and isinstance(job.field_selection, list):
            field_sel = job.field_selection  # type: ignore[assignment]

        data = format_rows(rows, job.format, field_sel)

        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        ext = file_extension(job.format)
        out_path = EXPORTS_DIR / f"{job_id}.{ext}"
        out_path.write_bytes(data)

        job.status = "complete"
        job.file_path = str(out_path)
        job.row_count = len(rows)
        job.completed_at = datetime.now(timezone.utc)
        job.expires_at = datetime.now(timezone.utc) + timedelta(hours=EXPORT_TTL_HOURS)
        session.commit()
        log.info("Export job %s complete: %d rows, %d bytes", job_id, len(rows), len(data))

    except Exception as exc:
        log.exception("Export job %s failed", job_id)
        job.status = "failed"
        job.error_message = str(exc)
        job.completed_at = datetime.now(timezone.utc)
        session.commit()
        raise
